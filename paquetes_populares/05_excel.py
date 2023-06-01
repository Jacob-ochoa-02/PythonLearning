import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")

# print(wb.sheetnames)
# print(wb["Hoja 1"])  # Acceder a una Hoja en particular
# hoja = wb.active    # Nos devuelve las hojas activas
# print(hoja)

# hoja = wb.active
# wb.create_sheet("Hoja 3")   # Crear una hoja con su respectivo nombre
# print(wb.sheetnames)
hoja = wb.active

hoja3 = wb["Hoja 3"]    # Selecciono una hoja especifica
hoja3.title = "Nuevo titulo"    # Le doy un nuevo titulo a una hoja

print(
    hoja.max_row,   # Nos indica la cantidad de filas de la hoja
    hoja.max_column,    # # Nos indica la cantidad de columnas de la hoja
)

celda = hoja["A1"]  # Queremos una celga
print(celda)    # Nos devuelve un objeto de tipo cell
print(celda.value)  # El valor de dicha celda
# celda.value = "Nombre completo" # Forma de cambiarle el valor a dicha celda

celda2 = hoja.cell(row=2, column=1)
print(celda2.value)
# Nos devuelve el valor de dicha base que elegimos en las filas
# Y columnas, empieza en base 1, no 0

# print(
#     celda2.value,   # valor de la celda
#     celda2.row, # fila de la celda
#     celda2.column,  # columna de la celda
#     celda2.coordinate   # coordenada de la celda
# )

for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        # print(fila, columna, celda.value)

columna = hoja["A"]
fila = hoja["1"]
print(columna)  # Se imprimen todos los elementos de esa columna
print(fila)  # Se imprimen todos los elementos de esa fila


hoja.append([1, 2, 3])
print(hoja.rows)    # Imprimimos las filas que tiene esta hoja
hoja.delete_cols(1, 1)
# Primero el indice de donde quiero empezar a eliminar
# El 2do es la cantidad de columnas que deseo eliminar
# Eliminar columnas, si no indicamos,
# eliminara por defecto el valor de 1
hoja.delete_rows(1, 1)
# Primero el indice de donde quiero empezar a eliminar
# El 2do es la cantidad de filas que deseo eliminar
# Eliminar filas, si no indicamos, eliminará por defecto el valor de 1

wb.save("nuevo_excel.xlsx")
